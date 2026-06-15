from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter


OUTPUT_FILE = "IoT_취업준비_마스터시트.xlsx"
MAX_ROWS = 1000


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="Malgun Gothic", size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Malgun Gothic", size=10, color="222222")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def style_sheet(ws: Worksheet, headers: list[str], widths: list[int], wrap_cols: set[int] | None = None) -> None:
    wrap_cols = wrap_cols or set()
    ws.append(headers)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{MAX_ROWS}"

    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(idx)].width = widths[idx - 1]

    ws.row_dimensions[1].height = 28

    for row in range(2, MAX_ROWS + 1):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(
                horizontal="left",
                vertical="top",
                wrap_text=(col in wrap_cols),
            )


def add_dropdown(ws: Worksheet, cell_range: str, values: list[str]) -> None:
    escaped = ",".join(values)
    dv = DataValidation(type="list", formula1=f'"{escaped}"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(cell_range)


def create_dashboard(ws: Worksheet) -> None:
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 26
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 2

    title_cell = ws["B2"]
    title_cell.value = "IoT 개발자 취업준비 대시보드"
    title_cell.font = Font(name="Malgun Gothic", size=16, bold=True, color="1F4E78")

    section_fill = PatternFill("solid", fgColor="E8F0FA")
    card_fill = PatternFill("solid", fgColor="F8FBFF")

    kpis = [
        ("조사한 기업 수", '=COUNTA(\'채용공고 분석\'!A2:A1000)'),
        ("지원 예정 기업 수", '=COUNTIF(\'채용공고 분석\'!L2:L1000,"지원예정")'),
        ("지원 완료 기업 수", '=COUNTIF(\'채용공고 분석\'!L2:L1000,"지원완료")'),
        ("서류 합격 수", '=COUNTIF(\'채용공고 분석\'!L2:L1000,"서류합격")'),
        ("면접 진행 수", '=COUNTIF(\'채용공고 분석\'!L2:L1000,"면접예정")'),
        ("최종 합격 수", '=COUNTIF(\'채용공고 분석\'!L2:L1000,"최종합격")'),
    ]

    start_row = 4
    for i, (label, formula) in enumerate(kpis):
        row = start_row + i
        label_cell = ws.cell(row=row, column=2, value=label)
        value_cell = ws.cell(row=row, column=3, value=formula)
        for c in (label_cell, value_cell):
            c.fill = card_fill
            c.border = THIN_BORDER
            c.font = BODY_FONT
            c.alignment = Alignment(vertical="center")
        label_cell.font = Font(name="Malgun Gothic", size=10, bold=True, color="1F4E78")
        value_cell.alignment = Alignment(horizontal="right", vertical="center")
        value_cell.number_format = "#,##0"

    right_metrics = [
        ("현재 보유 자격증 수", '=COUNTIF(\'자격증 및 공부계획\'!C2:C1000,"취득완료")', "#,##0"),
        ("진행 중 프로젝트 수", '=COUNTIFS(\'프로젝트 기술서\'!A2:A1000,"<>",\'프로젝트 기술서\'!P2:P1000,"X")', "#,##0"),
        ("자격증 평균 진행률", '=IFERROR(AVERAGE(\'자격증 및 공부계획\'!F2:F1000),0)', "0%"),
    ]

    for i, (label, formula, nfmt) in enumerate(right_metrics):
        row = 4 + i
        label_cell = ws.cell(row=row, column=4, value=label)
        value_cell = ws.cell(row=row, column=5, value=formula)
        for c in (label_cell, value_cell):
            c.fill = card_fill
            c.border = THIN_BORDER
            c.font = BODY_FONT
            c.alignment = Alignment(vertical="center")
        label_cell.font = Font(name="Malgun Gothic", size=10, bold=True, color="1F4E78")
        value_cell.alignment = Alignment(horizontal="right", vertical="center")
        value_cell.number_format = nfmt

    ws["B9"] = "이번 주 할 일"
    ws["B9"].font = Font(name="Malgun Gothic", size=12, bold=True, color="1F4E78")

    for col, text in enumerate(["구분", "항목", "마감일"], start=2):
        c = ws.cell(row=10, column=col, value=text)
        c.fill = section_fill
        c.font = Font(name="Malgun Gothic", size=10, bold=True, color="1F4E78")
        c.border = THIN_BORDER
        c.alignment = Alignment(horizontal="center")

    task_formula = (
        "=LET("
        "job,IFERROR(FILTER(HSTACK('채용공고 분석'!A2:A1000,'채용공고 분석'!K2:K1000),"
        "('채용공고 분석'!K2:K1000>=TODAY())*('채용공고 분석'!K2:K1000<=TODAY()+7)"
        "*('채용공고 분석'!L2:L1000<>\"최종합격\")*('채용공고 분석'!L2:L1000<>\"불합격\")),\"\"),"
        "cert,IFERROR(FILTER(HSTACK('자격증 및 공부계획'!B2:B1000,'자격증 및 공부계획'!D2:D1000),"
        "('자격증 및 공부계획'!D2:D1000>=TODAY())*('자격증 및 공부계획'!D2:D1000<=TODAY()+7)"
        "*('자격증 및 공부계획'!C2:C1000<>\"취득완료\")),\"\"),"
        "jobRows,IF(job=\"\",0,ROWS(job)),"
        "jobTagged,IF(jobRows=0,\"\",HSTACK(MAKEARRAY(jobRows,1,LAMBDA(r,c,\"채용\")),job)),"
        "certRows,IF(cert=\"\",0,ROWS(cert)),"
        "certTagged,IF(certRows=0,\"\",HSTACK(MAKEARRAY(certRows,1,LAMBDA(r,c,\"자격증\")),cert)),"
        "merged,VSTACK(jobTagged,certTagged),"
        "IF(merged=\"\",\"이번 주 마감 일정 없음\",TAKE(SORT(merged,3,1),8))"
        ")"
    )

    ws["B11"] = task_formula
    ws.merge_cells("B11:D18")
    ws["B11"].alignment = Alignment(vertical="top", horizontal="left", wrap_text=True)
    ws["B11"].border = THIN_BORDER
    ws["B11"].fill = PatternFill("solid", fgColor="FFFFFF")
    ws["B11"].font = BODY_FONT

    note = ws["B20"]
    note.value = "권장 흐름: 채용공고 분석 → 직무역량 정리 → 프로젝트 기술서 → 자소서 작성 → 면접 준비"
    note.font = Font(name="Malgun Gothic", size=10, italic=True, color="666666")


def create_workbook() -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)

    # 1) 대시보드
    dashboard = wb.create_sheet("대시보드")
    create_dashboard(dashboard)

    # 2) 채용공고 분석
    ws2 = wb.create_sheet("채용공고 분석")
    headers2 = [
        "기업명",
        "산업",
        "기업규모",
        "지역",
        "직무",
        "연봉",
        "필수조건",
        "우대사항",
        "기술스택",
        "채용공고 링크",
        "마감일",
        "지원상태",
        "관심도",
        "메모",
    ]
    widths2 = [18, 12, 12, 12, 14, 12, 24, 24, 18, 24, 12, 12, 10, 24]
    style_sheet(ws2, headers2, widths2, wrap_cols={7, 8, 14})
    ws2.auto_filter.ref = f"A1:N{MAX_ROWS}"

    add_dropdown(
        ws2,
        f"L2:L{MAX_ROWS}",
        ["조사중", "지원예정", "지원완료", "서류합격", "면접예정", "최종합격", "불합격"],
    )
    add_dropdown(ws2, f"M2:M{MAX_ROWS}", ["★★★★★", "★★★★", "★★★", "★★", "★"])

    date_style = "yyyy-mm-dd"
    for row in range(2, MAX_ROWS + 1):
        ws2[f"K{row}"].number_format = date_style

    range_row = f"A2:N{MAX_ROWS}"
    ws2.conditional_formatting.add(
        range_row,
        FormulaRule(formula=['$L2="최종합격"'], fill=PatternFill("solid", fgColor="D9EAD3")),
    )
    ws2.conditional_formatting.add(
        range_row,
        FormulaRule(formula=['$L2="지원완료"'], fill=PatternFill("solid", fgColor="D9E1F2")),
    )
    ws2.conditional_formatting.add(
        range_row,
        FormulaRule(
            formula=[
                '=AND($K2<>"",$K2>=TODAY(),$K2<=TODAY()+7,$L2<>"최종합격",$L2<>"불합격")'
            ],
            fill=PatternFill("solid", fgColor="FCE4D6"),
        ),
    )

    # 3) 직무역량 정리
    ws3 = wb.create_sheet("직무역량 정리")
    headers3 = ["구분", "항목명", "설명", "관련기술", "증빙자료", "직무연관성", "중요도", "비고"]
    widths3 = [10, 18, 26, 18, 18, 14, 10, 20]
    style_sheet(ws3, headers3, widths3, wrap_cols={3, 8})
    ws3.auto_filter.ref = f"A1:H{MAX_ROWS}"

    add_dropdown(ws3, f"A2:A{MAX_ROWS}", ["교육", "자격증", "기술", "프로젝트", "활동"])
    add_dropdown(ws3, f"F2:F{MAX_ROWS}", ["매우높음", "높음", "보통", "낮음"])
    dv_importance = DataValidation(type="whole", operator="between", formula1="1", formula2="5", allow_blank=True)
    ws3.add_data_validation(dv_importance)
    dv_importance.add(f"G2:G{MAX_ROWS}")

    # 4) 프로젝트 기술서
    ws4 = wb.create_sheet("프로젝트 기술서")
    headers4 = [
        "프로젝트명",
        "기간",
        "인원",
        "프로젝트 배경",
        "프로젝트 목적",
        "사용 기술",
        "시스템 구조",
        "나의 역할",
        "주요 기능",
        "문제점",
        "해결 방법",
        "결과",
        "배운 점",
        "직무연관성",
        "Github 링크",
        "포트폴리오 반영 여부",
    ]
    widths4 = [18, 14, 8, 24, 24, 16, 18, 18, 18, 20, 20, 18, 18, 12, 24, 14]
    style_sheet(ws4, headers4, widths4, wrap_cols={4, 5, 7, 8, 9, 10, 11, 12, 13})
    ws4.auto_filter.ref = f"A1:P{MAX_ROWS}"

    add_dropdown(ws4, f"P2:P{MAX_ROWS}", ["O", "X"])
    for row in range(2, MAX_ROWS + 1):
        ws4[f"B{row}"].number_format = date_style

    # 5) 자격증 및 공부계획
    ws5 = wb.create_sheet("자격증 및 공부계획")
    headers5 = ["구분", "목표", "현재상태", "시험일", "우선순위", "진행률", "예상완료일", "비고"]
    widths5 = [12, 20, 14, 12, 10, 10, 12, 24]
    style_sheet(ws5, headers5, widths5, wrap_cols={8})
    ws5.auto_filter.ref = f"A1:H{MAX_ROWS}"

    add_dropdown(ws5, f"C2:C{MAX_ROWS}", ["예정", "공부중", "접수완료", "취득완료"])
    add_dropdown(ws5, f"E2:E{MAX_ROWS}", ["★★★★★", "★★★★", "★★★", "★★", "★"])
    dv_progress = DataValidation(type="decimal", operator="between", formula1="0", formula2="1", allow_blank=True)
    ws5.add_data_validation(dv_progress)
    dv_progress.add(f"F2:F{MAX_ROWS}")

    for row in range(2, MAX_ROWS + 1):
        ws5[f"D{row}"].number_format = date_style
        ws5[f"G{row}"].number_format = date_style
        ws5[f"F{row}"].number_format = "0%"

    ws5.conditional_formatting.add(
        f"F2:F{MAX_ROWS}",
        CellIsRule(operator="equal", formula=["1"], fill=PatternFill("solid", fgColor="D9EAD3")),
    )
    ws5.conditional_formatting.add(
        f"F2:F{MAX_ROWS}",
        CellIsRule(
            operator="between",
            formula=["0.5", "0.9999"],
            fill=PatternFill("solid", fgColor="FFF2CC"),
        ),
    )
    ws5.conditional_formatting.add(
        f"F2:F{MAX_ROWS}",
        CellIsRule(operator="lessThan", formula=["0.5"], fill=PatternFill("solid", fgColor="F4CCCC")),
    )

    # 6) 자소서 재료 창고
    ws6 = wb.create_sheet("자소서 재료 창고")
    headers6 = ["경험명", "구분", "상황", "문제", "해결", "결과", "활용가능문항", "키워드"]
    widths6 = [18, 12, 22, 22, 22, 18, 24, 20]
    style_sheet(ws6, headers6, widths6, wrap_cols={3, 4, 5, 6, 7})
    ws6.auto_filter.ref = f"A1:H{MAX_ROWS}"

    # 7) 면접 질문 정리
    ws7 = wb.create_sheet("면접 질문 정리")
    headers7 = ["구분", "질문", "답변", "완성도", "최종수정일"]
    widths7 = [10, 30, 40, 14, 14]
    style_sheet(ws7, headers7, widths7, wrap_cols={2, 3})
    ws7.auto_filter.ref = f"A1:E{MAX_ROWS}"

    add_dropdown(ws7, f"A2:A{MAX_ROWS}", ["기업", "직무", "프로젝트", "인성"])
    add_dropdown(ws7, f"D2:D{MAX_ROWS}", ["미작성", "초안", "작성완료", "암기완료"])
    for row in range(2, MAX_ROWS + 1):
        ws7[f"E{row}"].number_format = date_style

    # 샘플 한 줄(초기 사용 편의)
    ws6.append(["졸업작품", "IoT 과정", "스마트홈 센서 대시보드 제작", "실시간 데이터 지연", "MQTT + 비동기 처리 도입", "응답속도 30% 개선", "경험 및 프로젝트", "IoT,WPF,MQTT"])
    for col in range(1, 9):
        c = ws6.cell(row=2, column=col)
        c.font = BODY_FONT
        c.border = THIN_BORDER
        c.alignment = Alignment(vertical="top", wrap_text=(col in {3, 4, 5, 6, 7}))

    return wb


def main() -> None:
    workbook = create_workbook()
    output = Path.cwd() / OUTPUT_FILE
    workbook.save(output)
    print(f"[완료] 파일 생성: {output}")
    print(f"[시트 수] {len(workbook.sheetnames)} / {', '.join(workbook.sheetnames)}")
    print(f"[생성 시각] {datetime.now().isoformat(timespec='seconds')}")


if __name__ == "__main__":
    main()
